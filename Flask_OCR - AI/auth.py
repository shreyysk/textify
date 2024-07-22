import secrets
import string
from werkzeug.security import check_password_hash, generate_password_hash
import re
import random
import openpyxl

PUNCTUATIONS = "@#$%&"
DEFAULT_PASSWORD_LENGTH = 12

def generate_password(length=DEFAULT_PASSWORD_LENGTH):
    characters = string.ascii_letters + string.digits + PUNCTUATIONS
    password = ''.join(secrets.choice(characters) for _ in range(length))
    return password

def register_user(username, email, password):
    workbook = openpyxl.load_workbook('users.xlsx')
    sheet = workbook.active

    # Check if the username already exists
    for row in sheet.iter_rows(values_only=True):
        if row[0] == username:
            return None, "Username already exists. Please try a different username."

    # Generate a hashed password
    hashed_password = generate_password_hash(password)

    # Add the new user to the Excel sheet
    sheet.append([username, email, hashed_password])
    workbook.save('users.xlsx')

    return username, password

def authenticate_user(username, password):
    workbook = openpyxl.load_workbook('users.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        if row[0] == username:
            stored_password = row[2]
            if check_password_hash(stored_password, password):
                return True
            else:
                return False

    return False 

def get_user_id_from_database(username):
    workbook = openpyxl.load_workbook('users.xlsx')
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True):
        if row[0] == username:
            user_id = row[0]  # Assuming the ID is in the second column
            return user_id

    return None